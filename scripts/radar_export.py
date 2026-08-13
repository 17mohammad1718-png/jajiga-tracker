#!/usr/bin/env python3
"""
radar_export.py — خروجی اکسل (xlsx) رادار رقبا به دسکتاپ
=================================================================
دو فایل زیبا روی دسکتاپ می‌سازد (دقیقاً همان چیزی که در داشبورد است):
    radar-calendar-YYYY-MM-DD.xlsx   — تب تقویم (روزهای آینده + متریک‌ها)
    radar-database-YYYY-MM-DD.xlsx   — تب دیتابیس (روزهای ثبت‌شده گذشته)

داده از JSONهای جاسازی‌شده در competitor-radar.html خوانده می‌شود.

استفاده:
    python scripts/radar_export.py
"""
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from radar_common import JM, g2j

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(SCRIPT_DIR)
HTML = os.path.join(ROOT, "competitor-radar.html")
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")

FONT = "B Nazanin"

STATUS_FILL = {
    "booked": "6B2B2B", "free": "1A4D7A", "half": "7B5A1E",
    "peak": "4A3B6B", "weekend": "155E63", "past": "16202F",
    "nodata": "0F172A",
}

THIN = Side(style="thin", color="3A4A5E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="132A47")
MONTH_FILL = PatternFill("solid", fgColor="1B3A5F")
TITLE_FILL = PatternFill("solid", fgColor="0F2A43")
SUB_FILL = PatternFill("solid", fgColor="0B1526")
OWN_FILL = PatternFill("solid", fgColor="0C2B24")


def load_payload(html, tag):
    m = re.search(rf"<script type='application/json' id='{tag}'>(.*?)</script>", html, re.S)
    if not m:
        sys.exit(f"{tag} در HTML پیدا نشد — اول build_radar_dashboard.py اجرا شود.")
    return json.loads(m.group(1))


def jm_span(dates):
    """بازه‌های ماه شمسی: [(start_col, count, 'مرداد 1405'), ...] — ستون اول روز = 2."""
    spans = []
    cur = 2
    for d in dates:
        y, m, _ = g2j(*map(int, d["g"].split("-")))
        if spans and spans[-1][2] == (y, m):
            spans[-1][1] += 1
        else:
            spans.append([cur, 1, (y, m)])
        cur += 1
    return [(s, c, f"{JM[m - 1]} {y}") for s, c, (y, m) in spans]


def style_sheet(ws, title, subtitle, dates, extra_cols):
    ncols = 1 + len(dates) + extra_cols
    last = get_column_letter(ncols)
    ws.sheet_view.rightToLeft = True
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value = title
    c.fill = TITLE_FILL
    c.font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]
    c.value = subtitle
    c.fill = SUB_FILL
    c.font = Font(name=FONT, size=11, color="94A3B8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    return 3  # ردیف شروع هدر ماه


def header_rows(ws, start_row, dates, extra_headers):
    spans = jm_span(dates)
    for s, c, name in spans:
        ws.merge_cells(start_row=start_row, start_column=s, end_row=start_row, end_column=s + c - 1)
        cell = ws.cell(row=start_row, column=s)
        cell.value = name
        cell.fill = MONTH_FILL
        cell.font = Font(name=FONT, size=11, bold=True, color="F8FAFC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22
    a = ws.cell(row=start_row, column=1, value="اتاق")
    a.fill = HEADER_FILL
    a.font = Font(name=FONT, size=11, bold=True, color="7DD3FC")
    a.alignment = Alignment(horizontal="center", vertical="center")
    a.border = BORDER
    dr = start_row + 1
    for i, d in enumerate(dates):
        cell = ws.cell(row=dr, column=2 + i, value=f"{d.get('wd', '')} {d['j']}".strip())
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONT, size=10, bold=True, color="7DD3FC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[dr].height = 20
    for j, h in enumerate(extra_headers):
        cell = ws.cell(row=dr, column=2 + len(dates) + j, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONT, size=10, bold=True, color="7DD3FC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    return dr


def data_rows(ws, start_row, rooms, dates, extra_getter=None):
    for i, r in enumerate(rooms):
        rn = start_row + i
        label = r["label"] + (" (من)" if r.get("own") else "")
        if r.get("village"):
            label += f"\n{r['village']}"
        cell = ws.cell(row=rn, column=1, value=label)
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.border = BORDER
        cell.font = Font(name=FONT, size=10, bold=bool(r.get("own")),
                         color="38BDF8" if r.get("own") else "E2E8F0")
        if r.get("own"):
            cell.fill = OWN_FILL
        ws.row_dimensions[rn].height = 28
        for j, d in enumerate(dates):
            c = r["cells"].get(d["g"])
            cell = ws.cell(row=rn, column=2 + j, value=c["t"] if c else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            fill = STATUS_FILL.get(c["s"]) if c else None
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name=FONT, size=9,
                             color="FFFFFF" if c and c["s"] != "nodata" else "64748B")
        if extra_getter:
            for j, v in enumerate(extra_getter(r)):
                col = 2 + len(dates) + j
                cell = ws.cell(row=rn, column=col, value=v)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER
                cell.font = Font(name=FONT, size=10, color="E2E8F0")
    return start_row + len(rooms) - 1


def finalize(ws, start_row, last_row, ncols, dates, legend_text):
    ws.column_dimensions["A"].width = 42
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    for i in range(len(dates), ncols - 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 11
    ws.freeze_panes = f"B{start_row}"
    lr = last_row + 2
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=ncols)
    c = ws.cell(row=lr, column=1, value=legend_text)
    c.font = Font(name=FONT, size=9, color="64748B")
    c.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ncols)}{last_row}"


def build_workbook(html):
    """از HTML داشبورد، یک Workbook با دو شیت (تقویم + دیتابیس) می‌سازد.
    خروجی: (wb, today_calendar, today_database)"""
    F = load_payload(html, "radarData")       # تقویم (آینده)
    P = load_payload(html, "radarPastData")   # دیتابیس (گذشته)
    wb = Workbook()

    ws = wb.active
    ws.title = "تقویم"
    start = style_sheet(ws, "🛰️ رادار رقبا — تقویم (روزهای آینده)",
                        f"{len(F['rooms'])} اتاق · به‌روزرسانی: {F['today']}", F["dates"], 6)
    hr = header_rows(ws, start, F["dates"], ["۳۰ روز", "۶۰ روز", "۹۰ روز", "میانگین شب", "تخفیف", "پیک"])
    lr = data_rows(ws, hr + 1, F["rooms"], F["dates"],
                   lambda r: [r["occ30"], r["occ60"], r["occ90"], r["avg"], r["discount"], r["peak"]])
    finalize(ws, start, lr, 1 + len(F["dates"]) + 6, F["dates"],
             "راهنما — پر: قرمز · خالی: آبی · نیمه‌پر: نارنجی · پیک: بنفش · آخر هفته: فیروزه‌ای · بدون داده: تیره")

    ws2 = wb.create_sheet("دیتابیس")
    start2 = style_sheet(ws2, "🛰️ رادار رقبا — دیتابیس (روزهای گذشته)",
                         f"{len(P['dates'])} روز ثبت‌شده · {len(P['rooms'])} اتاق · به‌روزرسانی: {P['today']}",
                         P["dates"], 0)
    hr2 = header_rows(ws2, start2, P["dates"], [])
    lr2 = data_rows(ws2, hr2 + 1, P["rooms"], P["dates"])
    finalize(ws2, start2, lr2, 1 + len(P["dates"]), P["dates"],
             "راهنما — پر: قرمز · خالی: آبی · نیمه‌پر: نارنجی · بدون داده: تیره")

    return wb, F["today"], P["today"]


def main():
    html = open(HTML, encoding="utf-8").read()
    wb, t1, t2 = build_workbook(html)
    out1 = os.path.join(DESKTOP, f"radar-calendar-{t1}.xlsx")
    out2 = os.path.join(DESKTOP, f"radar-database-{t2}.xlsx")
    wb.save(out1)
    wb.save(out2)
    print(f"Saved:\n  {out1}\n  {out2}")


if __name__ == "__main__":
    main()
